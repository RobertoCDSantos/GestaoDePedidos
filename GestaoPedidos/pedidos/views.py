# Create your views here.
from django.http import HttpResponse
from django.shortcuts import render
from django.db.models import Avg
from django.db.models import F, Avg
from django.utils import timezone
import pytz
from custom_users.models import CustomUser
from pedidosHorario.models import PedidoHorario
from pedidosOutros.tables import PedidosTable
from openpyxl import *
from pedidosUC.forms import PedidosDisciplinaForm
from pedidosOutros.models import PedidoOutros
from pedidosHorario.models import PedidoHorario
from pedidosSala.models import PedidoSala
from pedidosUC.models import PedidoUC
from .forms import AnoLetivoForm,SemestreForm,DisciplinaForm
from .models import AnoLetivo, Curso, Disciplina, Docentes, Edificio, Funcionario, Instituicao, Linha, Sala,Semestre,Pedido,DSD, TipoSala, Turma
from .tables import AnoLetivoTable,PedidoTable
from io import BytesIO
import base64
import pandas as pd
from tkinter import TclError
import matplotlib.pyplot as plt
from datetime import timedelta, datetime, date
import numpy as np
import matplotlib
from dateutil.parser import parse
import openpyxl
from .forms import ExcelUploadForm
from django.core.files.storage import FileSystemStorage
from django.db.models import Q
matplotlib.use('Agg')
plt.switch_backend('agg')
from django.forms import formset_factory,inlineformset_factory
from django.contrib.auth.decorators import permission_required
import os
import re

import os
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows




def remove_timezone(dt):
    if dt is not None:
        return dt.replace(tzinfo=None)
    return None



@permission_required('custom_users.funcionario_rights')
def export_pedidos_to_excel(request):
    selected_pedidos = request.POST.getlist('selected_pedidos')  # Get the selected pedido IDs
    selected_pedidos = list(map(int, selected_pedidos))
    todos = request.POST.get('is_closed')
    
    if todos != None:
        pedidos = Pedido.objects.select_subclasses(PedidoUC,PedidoHorario,PedidoOutros,PedidoSala)
    else:
        pedidos = Pedido.objects.select_subclasses(PedidoUC,PedidoHorario,PedidoOutros,PedidoSala).filter(id__in=selected_pedidos)  # Filter pedidos based on selected IDs

    print(pedidos)
    workbook = Workbook()
    worksheet = workbook.active

    # Write the headers of the Excel file (optional)
    headers = ['NrPedido', 'Titulo', 'Descricao', 'DataAlvo', 'DataAnalise', 'DataValidacao', 'DataCriacao', 'Estado', 'DataAlteracao', 'AnoLetivo', 'Docente', 'Funcionario']  # Replace with your desired field names
    
    
    worksheet.append(headers)

    for obj in pedidos:
        print(obj.titulo)
        print(obj.descricao)
        print()
        # Extract the fields you want to export from the object
        #dataanalise = remove_timezone(obj.dataanalise)
        #datavalidacao = remove_timezone(obj.datavalidacao)
        dataanalise = obj.dataanalise
        datavalidacao = obj.datavalidacao
        funcionario = obj.funcionario.codigo if obj.funcionario is not None else ''
        if isinstance(obj,PedidoUC):
            for l in obj.linha_set.all():
                data_row = [
                    obj.id, obj.titulo, obj.descricao, obj.dataalvo,
                    dataanalise, datavalidacao, obj.datacriacao,
                    obj.estado, obj.dataalteracao, obj.docente.codigo,
                    funcionario
                ]  # Replace with your desired field names
                data_row.append(l.descricao)
                data_row.append(l.uc.nome)
                worksheet.append(data_row)
        if isinstance(obj,PedidoHorario):
            for l in obj.linhahorario_set.all():
                data_row = [
                    obj.id, obj.titulo, obj.descricao, obj.dataalvo,
                    dataanalise, datavalidacao, obj.datacriacao,
                    obj.estado, obj.dataalteracao, obj.docente.codigo,
                    funcionario
                ]  # Replace with your desired field names
                data_row.append(l.tipodepedido)
                data_row.append(l.diadasemana)
                data_row.append(l.horainicio)
                data_row.append(l.horafim)
                data_row.append(l.datainicio)
                data_row.append(l.datafim)
                data_row.append(l.novadiadasemana)
                data_row.append(l.novahorainicio)
                data_row.append(l.novahorafim)
                data_row.append(l.novadatainicio)
                data_row.append(l.novadatafim)
                worksheet.append(data_row)
        if isinstance(obj,PedidoSala):
            for l in obj.linhasala_set.all():
                data_row = [
                    obj.id, obj.titulo, obj.descricao, obj.dataalvo,
                    dataanalise, datavalidacao, obj.datacriacao,
                    obj.estado, obj.dataalteracao, obj.docente.codigo,
                    funcionario
                ]  # Replace with your desired field names
                data_row.append(l.detalhe)
                data_row.append(l.sala.nsala)
                data_row.append(l.horainicio)
                data_row.append(l.horainicioantigo)
                data_row.append(l.horafimantigo)
                data_row.append(l.tipopedido)
                data_row.append(l.categoriatemporal)
                worksheet.append(data_row)
        if isinstance(obj,PedidoOutros):
            for l in obj.linhaoutros_set.all():
                data_row = [
                    obj.id, obj.titulo, obj.descricao, obj.dataalvo,
                    dataanalise, datavalidacao, obj.datacriacao,
                    obj.estado, obj.dataalteracao, obj.docente.codigo,
                    funcionario
                ]  # Replace with your desired field names
                data_row.append(l.descricao)
                worksheet.append(data_row)

    # Set the timezone to None for all datetime objects in the worksheet
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.value = cell.value.replace(tzinfo=None)

    # Create a response object with the appropriate content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=objects.xlsx'

    workbook.save(response)
    return response

@permission_required('custom_users.funcionario_rights')
def export_pedidos_to_excel1(request):
    selected_pedidos = request.POST.getlist('selected_pedidos')  # Get the selected pedido IDs
    selected_pedidos = list(map(int, selected_pedidos))


    pedidos = Pedido.objects.filter(id__in=selected_pedidos)  # Filter pedidos based on selected IDs
    
    workbook = Workbook()
    worksheet = workbook.active

    # Write the headers of the Excel file (optional)
    headers = ['NrPedido', 'Titulo', 'Descricao', 'DataAlvo', 'DataAnalise', 'DataValidacao', 'DataCriacao', 'Estado', 'DataAlteracao', 'AnoLetivo', 'Docente', 'Funcionario']  # Replace with your desired field names
    
    
    worksheet.append(headers)

    for obj in pedidos:
        # Extract the fields you want to export from the object
        #dataanalise = remove_timezone(obj.dataanalise)
        #datavalidacao = remove_timezone(obj.datavalidacao)
        dataanalise = obj.dataanalise
        datavalidacao = obj.datavalidacao
        funcionario = obj.funcionario.codigo if obj.funcionario is not None else ''
        data_row = [
            obj.id, obj.titulo, obj.descricao, obj.dataalvo,
            dataanalise, datavalidacao, obj.datacriacao,
            obj.estado, obj.dataalteracao, obj.docente.codigo,
            funcionario
        ]  # Replace with your desired field names
        worksheet.append(data_row)

    # Set the timezone to None for all datetime objects in the worksheet
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.value = cell.value.replace(tzinfo=None)

    # Create a response object with the appropriate content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=objects.xlsx'

    workbook.save(response)
    return response

@permission_required('custom_users.any_rights')
def index(request):
    user = request.user
    opcoes = request.GET.getlist('opcao')
    id_estado = request.GET.get('id_estado')
    pesquisa = request.GET.get('pesquisa')
    tipopedido = request.GET.get('tipopedido')
    pedidos = Pedido.objects.select_subclasses(PedidoUC, PedidoHorario, PedidoOutros, PedidoSala).order_by('datacriacao')

    if 'Titulo' in opcoes:
        pedidos = pedidos.filter(titulo__isnull=False)
    if 'Descricao' in opcoes:
        pedidos = pedidos.filter(descricao__isnull=False)
    if 'ID' in opcoes:
        pedidos = pedidos.filter(id__isnull=False)

    if pesquisa:
        filtros = Q()
        if 'Titulo' in opcoes:
            filtros |= Q(titulo__icontains=pesquisa)
        if 'Descricao' in opcoes:
            filtros |= Q(descricao__icontains=pesquisa)
        if id_estado != 'todos':
            pedidos = pedidos.filter(estado__iexact=id_estado)
        pedidos = pedidos.filter(filtros)
        pedidos = pedidos.order_by('datacriacao')
    elif id_estado != None and id_estado != 'todos':
        pedidos = pedidos.filter(estado__iexact=id_estado)

    if tipopedido == 'PedidoHorario':
        pedidos = [p for p in pedidos if isinstance(p, PedidoHorario)]
    elif tipopedido == 'PedidoSala':
        pedidos = [p for p in pedidos if isinstance(p, PedidoSala)]
    elif tipopedido == 'PedidoUC':
        pedidos = [p for p in pedidos if isinstance(p, PedidoUC)]
    elif tipopedido == 'PedidoOutros':
        pedidos = [p for p in pedidos if isinstance(p, PedidoOutros)]
    else:
        pedidos = [p for p in pedidos if isinstance(p, Pedido)]

    table = PedidoTable(pedidos)
    table.paginate(page=request.GET.get("page", 1), per_page=10)
    return render(request, 'index_pedidos.html',
                  {'pedidos': pedidos, 'table': table, 'name': pesquisa, 'opcao': opcoes, 'id_estado': id_estado, 'tipopedido': tipopedido})

def app(request):
    return render(request,'app.html')


########################
##     Ano Letivo     ##
########################

def render_index_anos(request,message,sucesso):
    anos = AnoLetivo.objects.all()
    table = AnoLetivoTable(anos)
    table.paginate(page=request.GET.get("page", 1), per_page=10)
    if not sucesso:
        colorBack = '#FF0000'
        color = '#FFFFFF'
        mensagem = message
        return render(request, 'index_anos.html', {'mensagem': mensagem, 'table': table, 'colorBack': colorBack , 'color': color})
    else:
        colorBack = '#d4edda'
        color = '#155724'
        mensagem = message
        return render(request, 'index_anos.html', {'mensagem': mensagem, 'table': table, 'colorBack': colorBack , 'color': color})

@permission_required('custom_users.funcionario_rights')
def criar_ano_letivo(request):
    if request.method == 'POST':
        # Verificar datas e nome iguais
        d_inicio = request.POST['datainicio']
        d_fim = request.POST['datafim']
        a = request.POST['ano']
        anos_filtered_datas = AnoLetivo.objects.filter(datainicio=d_inicio).filter(datafim=d_fim)
        anos_filtered_nome = AnoLetivo.objects.filter(ano=a)
        form = AnoLetivoForm(request.POST)
        if form.is_valid() and len(anos_filtered_datas) == 0 and len(anos_filtered_nome) == 0:
            datainicio = form.cleaned_data['datainicio']
            datafim = form.cleaned_data['datafim']
            # Check overlaping
            if AnoLetivo.objects.filter(datainicio__lte=datafim, datafim__gte=datainicio).exists():
                return render_index_anos(request,"Anos letivos com datas intercaladas",False)
            else:
                ano_letivo = form.save(commit=False)
                ano_letivo.estado = "inativo"
                ano_letivo.save()
                return render_index_anos(request,"Anos letivo criado com sucesso",True)  # Redirect to a success page or another view
        else:
            colorBack = '#FF0000'
            color = '#FFFFFF'
            mensagem = "Ano letivo existente"
            return render(request,'criar_ano_letivo.html',{
                'mensagem': mensagem, 
                'colorBack': colorBack , 
                'color': color
            })
    else:
        form = AnoLetivoForm()
    return render(request, 'criar_ano_letivo.html', {'form': form})

@permission_required('custom_users.funcionario_rights')
def ativar_ano_letivo(request,ano_id):
    ano = AnoLetivo.objects.get(id=ano_id)
    anos = AnoLetivo.objects.all()
    for a in anos:
        a.estado = "inativo"
        a.save()
    ano.estado = "ativo"
    ano.save()
    return render_index_anos(request,"Ano letivo ativado",True)


@permission_required('custom_users.funcionario_rights')
def editar_ano_letivo(request, ano_id):
    ano = AnoLetivo.objects.get(id=ano_id)
    anos = AnoLetivo.objects.all()
    table = AnoLetivoTable(anos)
    table.paginate(page=request.GET.get("page", 1), per_page=10)

    colorBack = '#FF0000'
    color = '#FFFFFF'
    context = {
        'ano': ano.ano,
        'datainicio': ano.datainicio.strftime('%Y-%m-%d') if ano.datainicio else '',
        'datafim': ano.datafim.strftime('%Y-%m-%d') if ano.datafim else '',
        'mensagem': "",
        'colorBack': colorBack,
        'color': color,
        'anos': anos,
        'table': table
    }
    if request.method == 'POST':
        datainicio = datetime.strptime(request.POST.get('datainicio'), '%Y-%m-%d')
        datafim = datetime.strptime(request.POST.get('datafim'), '%Y-%m-%d')
        if (AnoLetivo.objects.filter(ano=ano.ano).count() >= 2):
            context['mensagem'] = "Já existe um Ano Letivo com o ano " + str(ano.ano)      
            return render(request, 'editar_ano_letivo.html', context)
        ano.datainicio = request.POST['datainicio']
        ano.datafim = request.POST['datafim']
        # Check overlaping
        if len(AnoLetivo.objects.filter(datainicio__lte=ano.datainicio, datafim__gte=ano.datafim)) >= 2:
            return render_index_anos(request,"Anos letivos com datas intercaladas",False)
        ano.save()
        context['mensagem'] = "Ano letivo editado com sucesso"
        context['colorBack'] = '#d4edda'
        context['color'] = '#155724'
        context['anos'] = anos
        context['table'] = table
        return render(request, 'index_anos.html', context)
    else:
        context = {
            'ano': ano.ano,
            'datainicio': ano.datainicio.strftime('%Y-%m-%d') if ano.datainicio else '',
            'datafim': ano.datafim.strftime('%Y-%m-%d') if ano.datafim else '',
        }
        return render(request, 'editar_ano_letivo.html', context)

@permission_required('custom_users.any_rights')
def listar_ano_letivo(request):
    anos = AnoLetivo.objects.all()
    table = AnoLetivoTable(anos)
    table.paginate(page=request.GET.get("page", 1), per_page=10)
    return render(request,'index_anos.html',{'anos': anos, 'table': table})

@permission_required('custom_users.funcionario_rights')
def eliminar_ano_letivo(request,ano_id):
    ano = AnoLetivo.objects.get(pk=ano_id)
    Pedido.objects.filter(anoletivo=ano).exists()
    if Pedido.objects.filter(anoletivo=ano).exists():
        return render_index_anos(request, 'Ano letivo com pedidos associados', False)
    if ano.estado == "ativo":
        return render_index_anos(request, 'Ano letivo está ativo!', False)
    else:
        AnoLetivo.objects.filter(id=ano_id).delete()
        return render_index_anos(request, 'Ano letivo eliminado com sucesso!', False)


########################
##      Semestre      ##
########################

@permission_required('custom_users.funcionario_rights')
def criar_semestre(request):
    if request.method == 'POST':
        semestre_form = SemestreForm(request.POST)
        if(semestre_form.is_valid()):
            semestre_form.save()
            return render(request,'app.html')
    SemestreForm()
    return render(request,'criar_semestre.html')


########################
##       Disciplina   ##
########################
@permission_required('custom_users.funcionario_rights')
def criar_disciplina(request):
    semestres  = Semestre.objects.all()
    if request.method == 'POST':
        diciplina_form = DisciplinaForm(request.POST)
        if(diciplina_form.is_valid()):
            uc = diciplina_form.save(commit=False)
            uc.estado = "ativa"
            uc.save()
            return render(request,'app.html')
    DisciplinaForm()
    return render(request,'criar_disciplina.html',{'semestres': semestres})



########################
##     Importar       ##
########################

@permission_required('custom_users.funcionario_rights')
def importar_ruc(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        file_extension = myfile.name.split('.')[-1]

        if file_extension != 'xls':
            colorBack = '#FF0000'
            color = '#FFFFFF'
            mensagem = "Ficheiro não é do formato xls"
            return render(request, 'importar_ruc.html', { 'mensagem': mensagem, 'colorBack': colorBack, 'color': color})

        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        uploaded_file_url = fs.url(filename)
        empexceldata = pd.read_excel(filename)

        print(empexceldata)

        for _, row in empexceldata.iterrows():
            if row['Docente']:
                arr = row['Docente'].split(' - ')
                print(arr[0])
                print(arr[1])
                try:
                    docente = Docentes.objects.get(codigo=arr[0])
                except Docentes.DoesNotExist:
                    docente = Docentes.objects.create(codigo=arr[0], nome=arr[1])

            conteudo_entre_parenteses = row['Tipo'].split("(")[1].split(")")[0]

            if conteudo_entre_parenteses == 'Curso':
                padrao = r"\((\d+)\)"
                resultado = re.search(padrao, row['Regência'])
                string = row['Regência'][:resultado.start()].strip()
                numero = resultado.group(1)

                if numero:
                    try:
                        curso = Curso.objects.get(codigo=numero, curso=string)
                    except Curso.DoesNotExist:
                        curso = Curso.objects.create(codigo=numero, curso=string)

            if conteudo_entre_parenteses == 'Disciplina':
                padrao = r"\((\d+)\)"
                resultado = re.search(padrao, row['Regência'])
                string = row['Regência'][:resultado.start()].strip()
                numero = resultado.group(1)
                semestre = Semestre.objects.get(id=1)
                try:
                    disciplina = Disciplina.objects.get(
                        nome=string,
                        codigodisciplina=numero,
                        semestreid=semestre
                    )
                except Disciplina.DoesNotExist:
                    disciplina = Disciplina.objects.create(
                        nome=string,
                        codigodisciplina=numero,
                        semestreid=semestre
                    )

        colorBack = '#d4edda'
        color = '#155724'
        mensagem = "Ficheiro RUC importado com sucesso"
        return render(request, 'importar_ruc.html', {
            'uploaded_file_url': uploaded_file_url, 'mensagem': mensagem, 'colorBack': colorBack, 'color': color })
    return render(request, 'importar_ruc.html')

@permission_required('custom_users.funcionario_rights')
def import_dsd(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        file_extension = myfile.name.split('.')[-1]

        if file_extension != 'xls':
            colorBack = '#FF0000'
            color = '#FFFFFF'
            mensagem = "Ficheiro não é do formato xls"
            return render(request, 'import_excel_db.html', { 'mensagem': mensagem, 'colorBack': colorBack, 'color': color})

        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        uploaded_file_url = fs.url(filename)
        empexceldata = pd.read_excel(filename)

        required_fields = ['Data início', 'Data fim', 'Horas serviço', 'Disciplina', 'Cód. disciplina', 'Turma',
                           'Horas semanais', 'Horas período', 'Código curso', 'Curso', 'Cód. Docente', 'Nome Docente',
                           'Período', 'Factor', 'Agrupamento']

        if not all(field in empexceldata.columns for field in required_fields):
            colorBack = '#FF0000'
            color = '#FFFFFF'
            mensagem = "Ficheiro não contém todos os campos necessários"
            return render(request, 'import_excel_db.html',
                          {'mensagem': mensagem, 'colorBack': colorBack, 'color': color})

        for _, row in empexceldata.iterrows():
            try:
                semestre = Semestre.objects.get(
                    inicio_data=datetime.strptime(row['Data início'], "%d/%m/%Y").date(),
                    fim_data=datetime.strptime(row['Data fim'], "%d/%m/%Y").date(),
                    nr_horas_service=row['Horas serviço'],
                )
            except Semestre.DoesNotExist:
                semestre = Semestre.objects.create(
                    inicio_data=datetime.strptime(row['Data início'], "%d/%m/%Y").date(),
                    fim_data=datetime.strptime(row['Data fim'], "%d/%m/%Y").date(),
                    nr_horas_service=row['Horas serviço'],
                )

            try:
                disciplina = Disciplina.objects.get(
                    nome=row['Disciplina'],
                    codigodisciplina=row['Cód. disciplina'],
                    semestreid=semestre
                )
            except Disciplina.DoesNotExist:
                disciplina = Disciplina.objects.create(
                    nome=row['Disciplina'],
                    codigodisciplina=row['Cód. disciplina'],
                    semestreid=semestre
                )

            try:
                turma = Turma.objects.get(
                    nturma=row['Turma'],
                    horas_semanais=row['Horas semanais'],
                    horas_periodo=row['Horas período']
                )
            except Turma.DoesNotExist:
                turma = Turma.objects.create(
                    nturma=row['Turma'],
                    horas_semanais=row['Horas semanais'],
                    horas_periodo=row['Horas período']
                )

            if row['Código curso'] is None:
                try:
                    curso = Curso.objects.get(codigo=row['Código curso'], curso=row['Curso'])
                except Curso.DoesNotExist:
                    curso = Curso.objects.create(codigo=row['Código curso'], curso=row['Curso'])



            try:
                docente = Docentes.objects.get(codigo=row['Cód. Docente'])
            except Docentes.DoesNotExist:
                docente = Docentes.objects.create(codigo=row['Cód. Docente'], nome=row['Nome Docente'])

            dsd_save = DSD.objects.create(
                periodo=row['Período'],
                codigo_curso=row['Código curso'],
                curso=row['Curso'],
                func_docente=row['Função docente'],
                inst_docente=row['Inst. docente'],
                depart_docente=row['Depart. docente'],
                factor=row['Factor'],
                nome_docente=row['Nome Docente'],
                agrupamento=row['Agrupamento']
            )
        dsd_save.save()
        colorBack = '#d4edda'
        color = '#155724'
        mensagem = "Ficheiro DSD importado com sucesso"
        return render(request, 'import_excel_db.html', {
            'uploaded_file_url': uploaded_file_url, 'mensagem': mensagem, 'colorBack': colorBack, 'color': color })
    return render(request, 'import_excel_db.html')



@permission_required('custom_users.funcionario_rights')
def upload_docentes1(request):
    print("Hello")
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb.active
            i = 0
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if i != 0:
                    docente = Docentes(
                        codigo=row[0],
                        ativo=row[2],
                        nome=row[3],
                        individuo=row[4],
                        data_de_nascimento=row[5],
                        sexo=row[6],
                        tipo_de_identificacao=row[7],
                        identificacao=row[8],
                        data_de_emissao_de_idenficacao=row[9],
                        nacionalidade=row[10],
                        arquivo=row[11],
                        data_de_validade_da_identificacao=row[12],
                        nif=row[13],
                        pais_fiscal=row[14],
                        digito_verificacao=row[15]
                    )
                    docente.save()
                i += 1
            colorBack = '#d4edda'
            color = '#155724'
            mensagem = "Ficheiro importado com sucesso"
            form = ExcelUploadForm()
            return render(request, 'upload_docente.html', {'form': form, 'colorBack':colorBack,'color': color ,'mensagem':mensagem})
        else:
            print(form.errors)
    else:
        form = ExcelUploadForm()
    return render(request, 'upload_docente.html', {'form': form})

@permission_required('custom_users.funcionario_rights')
def upload_docentes(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb.active
            headers = [cell.value for cell in worksheet[1]]
            print(headers)
            required_fields = ['Código', 'Docente', 'Ativo', 'Nome', 'Indivíduo', 'Data de nascimento', 'Sexo', 'Tipo de identificação', 'Identificação', 'Data de emissão da identificação', 'Nacionalidade', 'Arquivo', 'Data de validade da identificação', 'NIF', 'País fiscal', 'Digito verificação']

            if not all(field in headers for field in required_fields):
                print("MISSING")
                colorBack = '#FF0000'
                color = '#FFFFFF'
                mensagem = "Campos necessários em falta"
                form = ExcelUploadForm()
                return render(request, 'upload_docente.html', {'form': form, 'colorBack':colorBack,'color': color ,'mensagem':mensagem})

            i = 0
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if i != 0:
                    docente = Docentes(
                        codigo=row[1],
                        ativo=row[3],
                        nome=row[4],
                        individuo=row[5],
                        data_de_nascimento=row[6],
                        sexo=row[7],
                        tipo_de_identificacao=row[8],
                        identificacao=row[9],
                        data_de_emissao_de_idenficacao=row[10],
                        nacionalidade=row[11],
                        arquivo=row[12],
                        data_de_validade_da_identificacao=row[13],
                        nif=row[14],
                        pais_fiscal=row[15],
                        digito_verificacao=row[16]
                    )
                    docente.save()
                i += 1

            colorBack = '#d4edda'
            color = '#155724'
            mensagem = "Ficheiro importado com sucesso"
            form = ExcelUploadForm()
            return render(request, 'upload_docente.html', {'form': form, 'colorBack':colorBack,'color': color ,'mensagem':mensagem})
        else:
            print(form.errors)
    else:
        form = ExcelUploadForm()
    return render(request, 'upload_docente.html', {'form': form})

########################
##     Estatisticas   ##
########################

@permission_required('custom_users.any_rights')
def estatisticas1(request):
    pedidos = Pedido.objects.all().order_by('datacriacao')
    return render(request,'estatisticas1.html')

def estatisticas2(request):
    pedidos = Pedido.objects.all().order_by('datacriacao')
    return render(request,'estatisticas2.html')

def estatisticas3(request):
    pedidos = Pedido.objects.all().order_by('datacriacao')
    return render(request,'estatisticas3.html')

@permission_required('custom_users.funcionario_rights')
def pedidos_processados(request):
    if request.method == 'GET':
        data_inicio_str = request.GET.get('data_inicio')
        data_fim_str = request.GET.get('data_fim')
        tipopedido = request.GET.get('tipopedido')

        if data_inicio_str > data_fim_str:
            colorBack = '#FF0000'
            color = '#FFFFFF'
            mensagem = "A data do inicio não pode ser maior que a data do fim!"
            context = {
                'mensagem': mensagem,
                'colorBack': colorBack,
                'color': color
            }
            return render(request, 'estatisticas1.html', context)


        now = datetime.now()
        if data_inicio_str:
            data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d')
            if now < data_inicio:
                colorBack = '#FF0000'
                color = '#FFFFFF'
                mensagem = "A data do inicio não pode ser superior a do dia de hoje!"
                context = {
                    'mensagem': mensagem,
                    'colorBack': colorBack,
                    'color': color
                }
                return render(request, 'estatisticas1.html', context)

        pedidos = Pedido.objects.select_subclasses(PedidoUC, PedidoHorario, PedidoOutros, PedidoSala).order_by('datacriacao')
        funcionario = Funcionario.objects.get(customuser_ptr=request.user)
        pedidos = [p for p in pedidos if p.funcionario == funcionario]
        print(pedidos)
        if tipopedido == 'PedidoHorario':
            pedidos = [p for p in pedidos if isinstance(p, PedidoHorario)]
        elif tipopedido == 'PedidoSala':
            pedidos = [p for p in pedidos if isinstance(p, PedidoSala)]
        elif tipopedido == 'PedidoUC':
            pedidos = [p for p in pedidos if isinstance(p, PedidoUC)]
        elif tipopedido == 'PedidoOutros':
            pedidos = [p for p in pedidos if isinstance(p, PedidoOutros)]
        else:
            pedidos = [p for p in pedidos if isinstance(p, Pedido)]

        default = False
        print(pedidos)
        if not data_inicio_str or not data_fim_str:
            default = True
            dias = int(request.GET.get('dias', 300000))
            data_inicio = timezone.now() - timedelta(days=dias)
            data_fim = timezone.now()
        else:
            data_inicio = timezone.make_aware(datetime.strptime(data_inicio_str, '%Y-%m-%d'))
            data_fim = timezone.make_aware(datetime.strptime(data_fim_str, '%Y-%m-%d')) + timedelta(days=1)


        pedidos = [p for p in pedidos if data_inicio <= p.datacriacao <= data_fim]

        print(pedidos)
        estados = list(map(lambda p: p.estado, pedidos))
        estados_counts = pd.Series(estados).value_counts()

        if not pedidos:
            colorBack = '#FF0000'
            color = '#FFFFFF'
            mensagem = "Não tens pedidos suficientes para poderes processar essa estatistica!"
            context = {
                'mensagem': mensagem,
                'colorBack': colorBack,
                'color': color
            }
            return render(request, 'estatisticas1.html', context)

        try:
            fig, ax = plt.subplots()
            colors = {'validado': 'green', 'recusado': 'red', 'analise': 'yellow'}
            pie_colors = [colors.get(label) for label in estados_counts.index]
            ax.pie(estados_counts.values, labels=estados_counts.index, autopct='%1.1f%%', colors=pie_colors)
            if (data_inicio == data_fim - timedelta(days=1)):
                ax.set_title('Pedidos processados em {}'.format(data_inicio.strftime('%d/%m/%Y')))
            else:
                if (default):
                    ax.set_title('Pedidos processados no período desde o início até {}.'.format(data_fim.strftime('%d/%m/%Y')))
                else:
                    ax.set_title('Pedidos processados no período de {} a {}.'.format(data_inicio.strftime('%d/%m/%Y'), data_fim.strftime('%d/%m/%Y')))

            buffer = BytesIO()
            plt.savefig(buffer, format='png')
            buffer.seek(0)
            image_data = base64.b64encode(buffer.getvalue()).decode()
            plt.close()
            context = {'image_data': image_data}

            return render(request, 'estatisticas1.html', context)

        except TclError:
            pass


@permission_required('custom_users.any_rights')
def media_pedidos_processados(request):
    if request.method == 'GET':
        data_inicio = request.GET.get('data_inicio')
        data_fim = request.GET.get('data_fim')
        estado = request.GET.get('estado', 'todos')
        default = False
        if not data_inicio or not data_fim:
            default = True
            dias = int(request.GET.get('dias', 300000))
            data_inicio = timezone.now() - timedelta(days=dias)
            data_fim = timezone.now()
        else:
            data_inicio = datetime.strptime(data_inicio, "%Y-%m-%d")
            data_fim = datetime.strptime(data_fim, "%Y-%m-%d") + timedelta(days=1)

        if estado == 'todos':
            pedidos = Pedido.objects.filter(datacriacao__range=[data_inicio, data_fim])
        else:
            pedidos = Pedido.objects.filter(datacriacao__range=[data_inicio, data_fim], estado=estado)

        media = pedidos.aggregate(media_tempo=Avg(F('dataalteracao') - F('datacriacao')))
        media_tempo_validacao = str(media['media_tempo']).split(".")[0]  # "0:58:53"
        horas, minutos, segundos = media_tempo_validacao.split(":")
        if (default):
            media_tempo_validacao = "O tempo médio que levaram os pedidos do tipo '" + str(estado) + "' a serem processados é de " + horas + " horas " + minutos + " minutos e " + segundos + " segundos, desde o inicio até " + str(data_fim.date()) + "."
        else:
            media_tempo_validacao = "O tempo médio que levaram os pedidos do tipo '" + str(estado) + "' a serem processados é de " + horas + " horas " + minutos + " minutos e " + segundos + " segundos, desde " + str(data_inicio.date()) + " até " + str(data_fim.date()) + "."

        context = {
            'media_tempo_validacao': media_tempo_validacao
        }

        return render(request, 'estatisticas2.html', context)

@permission_required('custom_users.any_rights')
def numero_pedidos_processados(request):
    if request.method == 'GET':
        data_inicio = request.GET.get('data_inicio')
        data_fim = request.GET.get('data_fim')
        estado = request.GET.get('estado', 'todos')
        default = False
        if not data_inicio or not data_fim:
            default = True
            dias = int(request.GET.get('dias', 300000))
            data_inicio = timezone.now() - timedelta(days=dias)
            data_fim = timezone.now()
        else:
            data_inicio = datetime.strptime(data_inicio, "%Y-%m-%d")
            data_fim = datetime.strptime(data_fim, "%Y-%m-%d") + timedelta(days=1)

        if estado == 'todos':
            pedidos = Pedido.objects.filter(datacriacao__range=[data_inicio, data_fim])
        else:
            pedidos = Pedido.objects.filter(datacriacao__range=[data_inicio, data_fim], estado=estado)

        num_pedidos_processados = pedidos.count()
        if (default):
            media_tempo_validacao = "Foram processados " + str(num_pedidos_processados) + " pedidos do tipo '" + str(estado) + "' desde o inicio até " + str(data_fim.date()) + "."
        else:
            media_tempo_validacao = "Foram processados " + str(num_pedidos_processados) + " pedidos do tipo '" + str(estado) + "' de " + str(data_inicio.date())+ " até " + str(data_fim.date()) + "."

        context = {
            'numero_tempo_validacao': media_tempo_validacao
        }

        return render(request, 'estatisticas3.html', context)


@permission_required('custom_users.funcionario_rights')
def pedidos_processados_funcionario(request):
    colorback = '#FF0000'
    color = '#FFFFFF'
    context = {
        'colorBack': colorback,
        'color': color,
        'mensagem': ''
    }
    if request.method == 'GET':
        if not request.GET.get('data_inicio') or not request.GET.get('data_fim'):
            context['mensagem'] = "Nenhuma data escolhida!"
            return render(request, 'estatisticas.html', context)
            
        datainicio = datetime.strptime(request.GET.get('data_inicio'), '%Y-%m-%d')
        datafim = datetime.strptime(request.GET.get('data_fim'), '%Y-%m-%d')
        tipopedido = request.GET.get('tipopedido')
        if datainicio > datafim:
            context['mensagem'] = "A data de início não pode ser superior à data de fim!"
            return render(request, 'estatisticas.html', context)
        if datainicio > datetime.now():
            context['mensagem'] = "A data de início não pode ser superior à data de hoje!"
            return render(request, 'estatisticas.html', context)
        pedidos = Pedido.objects.select_subclasses(PedidoUC, PedidoSala, PedidoHorario, PedidoOutros)
        match tipopedido:
            case 'PedidoHorario':
                pedidos = [p for p in pedidos if isinstance(p, PedidoHorario)]
            case 'PedidoUC':
                pedidos = [p for p in pedidos if isinstance(p, PedidoUC)]
            case 'PedidoSala':
                pedidos = [p for p in pedidos if isinstance(p, PedidoSala)]
            case 'PedidoOutros':
                pedidos = [p for p in pedidos if isinstance(p, PedidoOutros)]
        
        pedidos = [p for p in pedidos if datainicio.date() <= p.datacriacao.date() <= datafim.date()]
        pedidos = [p for p in pedidos if p.estado == 'recusado' or p.estado == 'validado']
        
        if not pedidos:
            context['mensagem'] = "Não há pedidos suficientes para formar uma estatística"
            return render(request, 'estatisticas.html', context)
        
        context['numero_tempo_validacao'] = len(pedidos)
        return render(request, 'estatisticas.html', context)
        

########################
##        Pedidos     ##
########################

def render_index_pedidos(request,message,sucesso):
    pedidos = Pedido.objects.all()
    table = PedidoTable(pedidos)
    table.paginate(page=request.GET.get("page", 1), per_page=10)
    if not sucesso:
        colorBack = '#FF0000'
        color = '#FFFFFF'
        mensagem = message
        return render(request, 'index_pedidos.html', {'mensagem': mensagem, 'table': table, 'colorBack': colorBack , 'color': color})
    else:
        colorBack = '#d4edda'
        color = '#155724'
        mensagem = message
        return render(request, 'index_pedidos.html', {'mensagem': mensagem, 'table': table, 'colorBack': colorBack , 'color': color})

@permission_required('custom_users.funcionario_rights')
def validar_pedido(request,pedido_id):
    pedido = Pedido.objects.get(pk=pedido_id)
    if(pedido.estado == "analise"):
        if(request.user.codigo == pedido.funcionario.codigo):
            pedido.estado = "validado"
            pedido.datavalidacao = datetime.now()
            pedido.save()
            return render_index_pedidos(request,"O estado do pedido foi actualizado com sucesso, pelo funcionário",True)
        else:
            return render_index_pedidos(request,"O estado do pedido não pode ser actualizado por outro funcionario ",False)
    else:
        pedidos = Pedido.objects.all().order_by('datacriacao')
        table = PedidoTable(pedidos)
        table.paginate(page=request.GET.get("page", 1), per_page=10)
        return render(request, 'index_pedidos.html',{'table': table})
    
@permission_required('custom_users.funcionario_rights')
def rejeitar_pedido(request,pedido_id):
    pedido = Pedido.objects.get(pk=pedido_id)
    if(pedido.estado == "analise"):
        if(request.user.codigo == pedido.funcionario.codigo):
            pedido.estado = "recusado" 
            pedido.datavalidacao = datetime.now() 
            pedido.save()
            return render_index_pedidos(request,"O estado do pedido foi actualizado com sucesso, pelo funcionário",True)
        else:
            return render_index_pedidos(request,"O estado do pedido não pode ser actualizado por outro funcionario ",False)
    else:
        pedidos = Pedido.objects.all().order_by('datacriacao')
        table = PedidoTable(pedidos)
        table.paginate(page=request.GET.get("page", 1), per_page=10)
        return render(request, 'index_pedidos.html',{'table': table})
    
@permission_required('custom_users.funcionario_rights')
def associar_pedido(request, pedido_id):
    pedidos = Pedido.objects.all().order_by('datacriacao')
    table = PedidoTable(pedidos)
    table.paginate(page=request.GET.get("page", 1), per_page=10)

    pedido = Pedido.objects.get(id=pedido_id)

    if pedido.estado == 'espera':
        pedido.funcionario = Funcionario.objects.get(customuser_ptr=request.user)
        pedido.dataanalise = datetime.now()
        pedido.estado = "analise"
        pedido.save()
        colorBack = '#d4edda'
        color = '#155724'
        mensagem = "O pedido foi lhe associado com sucesso"
    else:
        if(request.user.codigo == pedido.funcionario.codigo):
            pedido.funcionarioid = None
            pedido.estado = "espera"
            pedido.dataanalise = None
            pedido.save()
            colorBack = '#d4edda'
            color = '#155724'
            mensagem = "O pedido foi desassociado de si com sucesso"
        else:
            colorBack = '#FF0000'
            color = '#FFFFFF'
            mensagem = "O pedido não pode ser desassociado por outro funcionario"
    pedido.save()
    return render(request, 'index_pedidos.html',  {'mensagem': mensagem, 'table': table, 'colorBack': colorBack, 'color': color})
    
    
@permission_required('custom_users.any_rights')
def view_pedido(request, pedido_id):
    pedido = Pedido.objects.get(pk=pedido_id)
    pedidos = Pedido.objects.select_subclasses(PedidoUC,PedidoHorario,PedidoSala,PedidoOutros).order_by('datacriacao')
    p = [p for p in pedidos if (p.id == pedido.id)][0]
    linhas = pedido.linha_set.all()
    if isinstance(p,PedidoUC):
        linhas_data = [{'id': linha.id, 'descricao': linha.descricao, 'uc': linha.uc} for linha in linhas]
        return render(request, 'view_pedido.html', {'pedido': p, 'linhas_data': linhas_data, 'tipo': 'uc'})
    if isinstance(p,PedidoHorario):
        return render(request, 'view_pedido.html', {'pedido': p, 'linhas_data': pedido.linhahorario_set.all(), 'tipo': 'horario'})
    if isinstance(p,PedidoOutros):
        return render(request, 'view_pedido.html', {'pedido': p, 'linhas_data': pedido.linhaoutros_set.all(), 'tipo': 'outros'})
    if isinstance(p,PedidoSala):
        return render(request, 'view_pedido.html', {'pedido': p, 'linhas_data': pedido.linhasala_set.all(), 'tipo': 'sala'})


def get_or_create_instituicao(dbframe):
    nome = str(dbframe['Nome Instituição'])
    if nome.lower() == 'nan' or nome  == '' or nome is None or isinstance(dbframe['Nome Instituição'], float):
        obj, created = Instituicao.objects.get_or_create(sigla='Ualg', nome='Universidade do Algarve')
    else:
        obj, created = Instituicao.objects.get_or_create(nome=dbframe['Nome Instituição'], sigla=(''.join([char for char in str(dbframe['Nome Instituição']) if char.isupper()])))
    if created:
        obj.save()
    return obj
    
def get_or_create_edificio(dbframe):
    match = re.search(r"^(.*?)\s*(\s*\([^)]*\))(?:\s*\(([^)]*)\))?$", dbframe['Desc. Edifício'])
    if match:
        nome = match.group(1)
        local = match.group(3) if match.group(3) else match.group(2).replace("(", "").replace(")", "")
        local = "Gambelas" if local == "Saúde" else local
    else:
        nome = dbframe['Desc. Edifício']
        temp = dbframe['Desc. Edifício'].split(" ")
        local = temp[len(temp) - 1]
    obj, created = Edificio.objects.get_or_create(descricao=str(dbframe['Desc. Edifício']), nome=nome, localizacao=local)
    if created:
        obj.instituicao.set([get_or_create_instituicao(dbframe)])
        obj.save()
    return obj

def get_or_create_tipos_sala(dbframe):
    types = dbframe['Id. tipo sala'].split(',')
    lista = []
    for choice in types:
        obj, created = TipoSala.objects.get_or_create(tipo=choice)
        if created:
            obj.save()
        lista.append(obj)
    return lista

@permission_required('custom_users.funcionario_rights')
def import_salas(request):
    if request.method == 'POST' and request.FILES['myfile']:      
        myfile = request.FILES['myfile']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        uploaded_file_url = fs.url(filename)
        empexceldata = pd.read_excel(filename)
        for _, dbframe in empexceldata.iterrows():
            ed = get_or_create_edificio(dbframe)
            nr_sala = dbframe['Desc. Sala'].split('-')[0]
            desc = dbframe['Des. Categoria']
            ano = AnoLetivo.objects.get(estado="ativo")
            obj = Sala.objects.create(nsala=nr_sala, lotacao=dbframe['Lotação presencial sala'], descricao=desc, edificioid=ed, anoletivo=ano)
            obj.tipo_sala.set(get_or_create_tipos_sala(dbframe))
            obj.save()
        colorBack = '#d4edda'
        color = '#155724'
        return render(request, 'import_sala.html', {'mensagem': "Importado com sucesso", 'colorBack': colorBack, 'color': color})
    else:
        return render(request, 'import_sala.html', {})